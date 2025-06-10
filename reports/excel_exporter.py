"""
Invoice Management System - Excel Report Generator
Advanced Excel report generation with multiple sheets, charts,
formatting, and business intelligence features.
"""

from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Dict, List, Optional, Any, Tuple
from pathlib import Path
import logging

# Excel libraries
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.formatting.rule import DataBarRule
from openpyxl.drawing import Image

from models.database import Invoice, Company, TkaWorker
from services.invoice_service import InvoiceService
from models.business import DataHelper, ReportHelper
from utils.formatters import (
    format_currency_idr, format_date_short, format_npwp_display,
    format_excel_currency, format_excel_date
)
from utils.helpers import ensure_directory, safe_filename
from config import export_config, app_config

logger = logging.getLogger(__name__)

class ExcelStyleManager:
    """Manages Excel styles and formatting"""
    
    def __init__(self, workbook: openpyxl.Workbook):
        self.workbook = workbook
        self._create_named_styles()
    
    def _create_named_styles(self):
        """Create named styles for consistent formatting"""
        # Header style
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_style.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_style.alignment = Alignment(horizontal='center', vertical='center')
        header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.workbook.add_named_style(header_style)
        
        # Subheader style
        subheader_style = NamedStyle(name="subheader_style")
        subheader_style.font = Font(name='Calibri', size=11, bold=True)
        subheader_style.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        subheader_style.alignment = Alignment(horizontal='center', vertical='center')
        subheader_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.workbook.add_named_style(subheader_style)
        
        # Data style
        data_style = NamedStyle(name="data_style")
        data_style.font = Font(name='Calibri', size=10)
        data_style.alignment = Alignment(horizontal='left', vertical='center')
        data_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.workbook.add_named_style(data_style)
        
        # Currency style
        currency_style = NamedStyle(name="currency_style")
        currency_style.font = Font(name='Calibri', size=10)
        currency_style.alignment = Alignment(horizontal='right', vertical='center')
        currency_style.number_format = '#,##0'
        currency_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.workbook.add_named_style(currency_style)
        
        # Date style
        date_style = NamedStyle(name="date_style")
        date_style.font = Font(name='Calibri', size=10)
        date_style.alignment = Alignment(horizontal='center', vertical='center')
        date_style.number_format = 'dd/mm/yyyy'
        date_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.workbook.add_named_style(date_style)
        
        # Title style
        title_style = NamedStyle(name="title_style")
        title_style.font = Font(name='Calibri', size=16, bold=True, color='366092')
        title_style.alignment = Alignment(horizontal='center', vertical='center')
        self.workbook.add_named_style(title_style)

class ExcelReportGenerator:
    """Main Excel report generator"""
    
    def __init__(self, session=None):
        self.session = session
        self.data_helper = DataHelper(session) if session else None
        self.report_helper = ReportHelper(session) if session else None
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.session:
            self.session.close()
    
    def generate_invoice_summary_report(self, start_date: date = None, end_date: date = None,
                                      company_ids: List[int] = None, output_path: str = None) -> str:
        """Generate comprehensive invoice summary report"""
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = safe_filename(f"Invoice_Summary_Report_{timestamp}.xlsx")
            output_dir = ensure_directory(export_config.export_directory)
            output_path = output_dir / filename
        
        # Create workbook
        wb = openpyxl.Workbook()
        style_manager = ExcelStyleManager(wb)
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Get data
        filters = {}
        if start_date:
            filters['start_date'] = start_date
        if end_date:
            filters['end_date'] = end_date
        if company_ids:
            filters['company_ids'] = company_ids
        
        invoice_data = self.report_helper.get_invoice_report_data(**filters) if self.report_helper else []
        
        # Create sheets
        self._create_summary_sheet(wb, invoice_data, start_date, end_date)
        self._create_detailed_sheet(wb, invoice_data)
        self._create_company_analysis_sheet(wb, invoice_data)
        self._create_monthly_trends_sheet(wb, invoice_data)
        self._create_status_analysis_sheet(wb, invoice_data)
        
        # Save workbook
        wb.save(output_path)
        
        logger.info(f"Generated invoice summary report: {output_path}")
        return str(output_path)
    
    def _create_summary_sheet(self, wb: openpyxl.Workbook, invoice_data: List[Dict], 
                            start_date: date = None, end_date: date = None):
        """Create executive summary sheet"""
        ws = wb.create_sheet("Executive Summary")
        
        # Title
        ws['A1'] = f"{app_config.name} - Invoice Summary Report"
        ws['A1'].style = "title_style"
        ws.merge_cells('A1:F1')
        
        # Report period
        period_text = "All Time"
        if start_date and end_date:
            period_text = f"{format_date_short(start_date)} - {format_date_short(end_date)}"
        elif start_date:
            period_text = f"From {format_date_short(start_date)}"
        elif end_date:
            period_text = f"Until {format_date_short(end_date)}"
        
        ws['A2'] = f"Period: {period_text}"
        ws['A2'].font = Font(name='Calibri', size=12, bold=True)
        ws.merge_cells('A2:F2')
        
        # Generate timestamp
        ws['A3'] = f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A3'].font = Font(name='Calibri', size=10, italic=True)
        ws.merge_cells('A3:F3')
        
        # Key metrics
        total_invoices = len(invoice_data)
        total_amount = sum(inv.get('total_amount', 0) for inv in invoice_data)
        avg_amount = total_amount / total_invoices if total_invoices > 0 else 0
        
        # Status breakdown
        status_counts = {}
        status_amounts = {}
        for inv in invoice_data:
            status = inv.get('status', 'unknown')
            status_counts[status] = status_counts.get(status, 0) + 1
            status_amounts[status] = status_amounts.get(status, 0) + inv.get('total_amount', 0)
        
        # Key metrics table
        metrics_start_row = 5
        
        # Headers
        headers = ['Metric', 'Value']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=metrics_start_row, column=col)
            cell.value = header
            cell.style = "header_style"
        
        # Metrics data
        metrics = [
            ['Total Invoices', total_invoices],
            ['Total Amount', total_amount],
            ['Average Invoice Amount', avg_amount],
            ['Draft Invoices', status_counts.get('draft', 0)],
            ['Finalized Invoices', status_counts.get('finalized', 0)],
            ['Paid Invoices', status_counts.get('paid', 0)]
        ]
        
        for row_idx, (metric, value) in enumerate(metrics, metrics_start_row + 1):
            ws.cell(row=row_idx, column=1, value=metric).style = "data_style"
            
            cell = ws.cell(row=row_idx, column=2)
            if 'Amount' in metric:
                cell.value = value
                cell.style = "currency_style"
            else:
                cell.value = value
                cell.style = "data_style"
        
        # Status breakdown chart would go here (simplified for now)
        chart_start_row = metrics_start_row + len(metrics) + 3
        
        ws.cell(row=chart_start_row, column=1, value="Status Breakdown").style = "subheader_style"
        
        status_headers = ['Status', 'Count', 'Amount']
        for col, header in enumerate(status_headers, 1):
            ws.cell(row=chart_start_row + 1, column=col).value = header
            ws.cell(row=chart_start_row + 1, column=col).style = "header_style"
        
        for row_idx, (status, count) in enumerate(status_counts.items(), chart_start_row + 2):
            ws.cell(row=row_idx, column=1, value=status.title()).style = "data_style"
            ws.cell(row=row_idx, column=2, value=count).style = "data_style"
            ws.cell(row=row_idx, column=3, value=status_amounts.get(status, 0)).style = "currency_style"
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_detailed_sheet(self, wb: openpyxl.Workbook, invoice_data: List[Dict]):
        """Create detailed invoice list sheet"""
        ws = wb.create_sheet("Invoice Details")
        
        # Headers
        headers = [
            'Invoice Number', 'Date', 'Company', 'NPWP', 'Status',
            'Subtotal', 'VAT Amount', 'Total Amount', 'Created By', 'Created Date'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.style = "header_style"
        
        # Write data
        for row, invoice in enumerate(invoice_data, 2):
            data = [
                invoice.get('invoice_number', ''),
                invoice.get('invoice_date'),
                invoice.get('company_name', ''),
                format_npwp_display(invoice.get('npwp', '')),
                invoice.get('status', '').title(),
                invoice.get('subtotal', 0),
                invoice.get('vat_amount', 0),
                invoice.get('total_amount', 0),
                invoice.get('creator_name', ''),
                invoice.get('created_at')
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                
                # Apply appropriate style
                if col in [6, 7, 8]:  # Currency columns
                    cell.style = "currency_style"
                elif col in [2, 10]:  # Date columns
                    cell.style = "date_style"
                else:
                    cell.style = "data_style"
        
        # Add data bars for amount columns
        if len(invoice_data) > 0:
            # Total amount data bar
            total_col = get_column_letter(8)
            data_range = f"{total_col}2:{total_col}{len(invoice_data) + 1}"
            rule = DataBarRule(start_type='min', end_type='max', color='5B9BD5')
            ws.conditional_formatting.add(data_range, rule)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
        
        # Freeze panes
        ws.freeze_panes = 'A2'
    
    def _create_company_analysis_sheet(self, wb: openpyxl.Workbook, invoice_data: List[Dict]):
        """Create company analysis sheet"""
        ws = wb.create_sheet("Company Analysis")
        
        # Group data by company
        company_stats = {}
        for invoice in invoice_data:
            company_name = invoice.get('company_name', 'Unknown')
            if company_name not in company_stats:
                company_stats[company_name] = {
                    'count': 0,
                    'total_amount': 0,
                    'avg_amount': 0,
                    'npwp': invoice.get('npwp', ''),
                    'statuses': {}
                }
            
            stats = company_stats[company_name]
            stats['count'] += 1
            stats['total_amount'] += invoice.get('total_amount', 0)
            
            status = invoice.get('status', 'unknown')
            stats['statuses'][status] = stats['statuses'].get(status, 0) + 1
        
        # Calculate averages
        for company, stats in company_stats.items():
            if stats['count'] > 0:
                stats['avg_amount'] = stats['total_amount'] / stats['count']
        
        # Sort by total amount
        sorted_companies = sorted(company_stats.items(), 
                                key=lambda x: x[1]['total_amount'], reverse=True)
        
        # Headers
        headers = [
            'Company Name', 'NPWP', 'Invoice Count', 'Total Amount',
            'Average Amount', 'Draft', 'Finalized', 'Paid'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.style = "header_style"
        
        # Data
        for row, (company_name, stats) in enumerate(sorted_companies, 2):
            data = [
                company_name,
                format_npwp_display(stats['npwp']),
                stats['count'],
                stats['total_amount'],
                stats['avg_amount'],
                stats['statuses'].get('draft', 0),
                stats['statuses'].get('finalized', 0),
                stats['statuses'].get('paid', 0)
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                
                if col in [4, 5]:  # Currency columns
                    cell.style = "currency_style"
                elif col in [3, 6, 7, 8]:  # Count columns
                    cell.style = "data_style"
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.style = "data_style"
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
        
        # Create chart if we have data
        if len(sorted_companies) > 0:
            self._create_company_chart(ws, len(sorted_companies))
    
    def _create_company_chart(self, ws, data_rows: int):
        """Create chart for company analysis"""
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Total Amount by Company"
        chart.y_axis.title = 'Amount (Rp)'
        chart.x_axis.title = 'Company'
        
        # Data for chart (company names and total amounts)
        data = Reference(ws, min_col=4, min_row=1, max_row=min(data_rows + 1, 21), max_col=4)  # Limit to top 20
        cats = Reference(ws, min_col=1, min_row=2, max_row=min(data_rows + 1, 21))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        
        # Add chart to worksheet
        ws.add_chart(chart, "J2")
    
    def _create_monthly_trends_sheet(self, wb: openpyxl.Workbook, invoice_data: List[Dict]):
        """Create monthly trends analysis sheet"""
        ws = wb.create_sheet("Monthly Trends")
        
        # Group data by month
        monthly_stats = {}
        for invoice in invoice_data:
            invoice_date = invoice.get('invoice_date')
            if isinstance(invoice_date, str):
                try:
                    invoice_date = datetime.strptime(invoice_date, '%Y-%m-%d').date()
                except ValueError:
                    continue
            elif not isinstance(invoice_date, date):
                continue
            
            month_key = f"{invoice_date.year}-{invoice_date.month:02d}"
            if month_key not in monthly_stats:
                monthly_stats[month_key] = {
                    'count': 0,
                    'total_amount': 0,
                    'year': invoice_date.year,
                    'month': invoice_date.month
                }
            
            monthly_stats[month_key]['count'] += 1
            monthly_stats[month_key]['total_amount'] += invoice.get('total_amount', 0)
        
        # Sort by month
        sorted_months = sorted(monthly_stats.items())
        
        # Headers
        headers = ['Year', 'Month', 'Invoice Count', 'Total Amount', 'Average Amount']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.style = "header_style"
        
        # Data
        month_names = [
            'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
            'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'
        ]
        
        for row, (month_key, stats) in enumerate(sorted_months, 2):
            avg_amount = stats['total_amount'] / stats['count'] if stats['count'] > 0 else 0
            
            data = [
                stats['year'],
                month_names[stats['month'] - 1],
                stats['count'],
                stats['total_amount'],
                avg_amount
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                
                if col in [4, 5]:  # Currency columns
                    cell.style = "currency_style"
                else:
                    cell.style = "data_style"
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
        
        # Create trend chart
        if len(sorted_months) > 1:
            self._create_trend_chart(ws, len(sorted_months))
    
    def _create_trend_chart(self, ws, data_rows: int):
        """Create trend chart for monthly analysis"""
        chart = LineChart()
        chart.title = "Monthly Invoice Trends"
        chart.style = 13
        chart.y_axis.title = 'Amount (Rp)'
        chart.x_axis.title = 'Month'
        
        # Data for chart
        data = Reference(ws, min_col=4, min_row=1, max_row=data_rows + 1, max_col=5)
        cats = Reference(ws, min_col=2, min_row=2, max_row=data_rows + 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Add chart to worksheet
        ws.add_chart(chart, "G2")
    
    def _create_status_analysis_sheet(self, wb: openpyxl.Workbook, invoice_data: List[Dict]):
        """Create status analysis sheet"""
        ws = wb.create_sheet("Status Analysis")
        
        # Status statistics
        status_stats = {}
        for invoice in invoice_data:
            status = invoice.get('status', 'unknown').title()
            if status not in status_stats:
                status_stats[status] = {
                    'count': 0,
                    'total_amount': 0,
                    'percentage': 0
                }
            
            status_stats[status]['count'] += 1
            status_stats[status]['total_amount'] += invoice.get('total_amount', 0)
        
        # Calculate percentages
        total_invoices = len(invoice_data)
        for status, stats in status_stats.items():
            if total_invoices > 0:
                stats['percentage'] = (stats['count'] / total_invoices) * 100
        
        # Headers
        headers = ['Status', 'Count', 'Percentage', 'Total Amount', 'Average Amount']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.style = "header_style"
        
        # Data
        for row, (status, stats) in enumerate(status_stats.items(), 2):
            avg_amount = stats['total_amount'] / stats['count'] if stats['count'] > 0 else 0
            
            data = [
                status,
                stats['count'],
                stats['percentage'],
                stats['total_amount'],
                avg_amount
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                
                if col == 3:  # Percentage
                    cell.style = "data_style"
                    cell.number_format = '0.00%'
                elif col in [4, 5]:  # Currency
                    cell.style = "currency_style"
                else:
                    cell.style = "data_style"
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
        
        # Create pie chart
        if len(status_stats) > 0:
            self._create_status_pie_chart(ws, len(status_stats))
    
    def _create_status_pie_chart(self, ws, data_rows: int):
        """Create pie chart for status analysis"""
        chart = PieChart()
        chart.title = "Invoice Status Distribution"
        
        # Data for chart
        data = Reference(ws, min_col=2, min_row=1, max_row=data_rows + 1, max_col=2)
        cats = Reference(ws, min_col=1, min_row=2, max_row=data_rows + 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Add chart to worksheet
        ws.add_chart(chart, "G2")
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

# Global instance
excel_generator = ExcelReportGenerator()

def generate_invoice_summary_report(start_date: date = None, end_date: date = None,
                                  company_ids: List[int] = None, output_path: str = None) -> str:
    """Convenience function to generate invoice summary report"""
    with ExcelReportGenerator() as generator:
        return generator.generate_invoice_summary_report(start_date, end_date, company_ids, output_path)

if __name__ == "__main__":
    # Test Excel generator
    print("Testing Excel generator...")
    
    try:
        # Test would require actual data
        # report_path = generate_invoice_summary_report()
        # print(f"Generated report: {report_path}")
        
        print("✅ Excel generator initialized successfully")
        
    except Exception as e:
        print(f"❌ Excel generator test failed: {e}")