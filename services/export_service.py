"""
Invoice Management System - Export Service
Comprehensive service for exporting invoices to PDF and Excel formats
with professional layouts and business formatting.
"""

import os
from pathlib import Path
from datetime import date, datetime
from decimal import Decimal
from typing import Dict, List, Optional, Any, Union, BinaryIO
import logging

# PDF Generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, 
    PageBreak, Frame, PageTemplate, BaseDocTemplate
)
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY

# Excel Generation
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing import Image as XLImage

from models.database import Invoice, get_db_session
from services.invoice_service import InvoiceService
from utils.formatters import (
    format_currency_idr, format_date_long, format_date_short,
    format_currency_words, format_npwp_display
)
from utils.helpers import ensure_directory, safe_filename
from config import export_config, app_config

logger = logging.getLogger(__name__)

class PDFExportService:
    """Service for exporting invoices to PDF format"""
    
    def __init__(self):
        self.page_size = letter  # Can be changed to A4 if needed
        self.margin = 0.75 * inch
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom paragraph styles"""
        # Header style
        self.styles.add(ParagraphStyle(
            'CustomHeader',
            parent=self.styles['Heading1'],
            fontSize=20,
            textColor=colors.black,
            alignment=TA_CENTER,
            spaceAfter=12
        ))
        
        # Company info style
        self.styles.add(ParagraphStyle(
            'CompanyInfo',
            parent=self.styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            spaceAfter=6
        ))
        
        # Invoice details style
        self.styles.add(ParagraphStyle(
            'InvoiceDetails',
            parent=self.styles['Normal'],
            fontSize=10,
            alignment=TA_LEFT,
            spaceAfter=6
        ))
        
        # Table header style
        self.styles.add(ParagraphStyle(
            'TableHeader',
            parent=self.styles['Normal'],
            fontSize=8,
            alignment=TA_CENTER,
            textColor=colors.black
        ))
        
        # Table cell style
        self.styles.add(ParagraphStyle(
            'TableCell',
            parent=self.styles['Normal'],
            fontSize=8,
            alignment=TA_LEFT
        ))
    
    def export_invoice_pdf(self, invoice_id: int, output_path: str = None) -> str:
        """
        Export single invoice to PDF
        
        Args:
            invoice_id: ID of invoice to export
            output_path: Optional custom output path
            
        Returns:
            Path to generated PDF file
        """
        with InvoiceService() as service:
            invoice = service.get_invoice(invoice_id)
            if not invoice:
                raise ValueError(f"Invoice {invoice_id} not found")
            
            # Generate filename if not provided
            if not output_path:
                filename = safe_filename(f"Invoice_{invoice.invoice_number}.pdf")
                output_dir = ensure_directory(export_config.export_directory)
                output_path = output_dir / filename
            
            # Create PDF document
            doc = SimpleDocTemplate(
                str(output_path),
                pagesize=self.page_size,
                rightMargin=self.margin,
                leftMargin=self.margin,
                topMargin=self.margin,
                bottomMargin=self.margin
            )
            
            # Build PDF content
            story = self._build_invoice_content(invoice)
            
            # Generate PDF
            doc.build(story)
            
            logger.info(f"Exported invoice {invoice.invoice_number} to PDF: {output_path}")
            return str(output_path)
    
    def _build_invoice_content(self, invoice: Invoice) -> List:
        """Build PDF content for invoice"""
        story = []
        
        # Company header
        story.extend(self._build_header())
        
        # Invoice details
        story.extend(self._build_invoice_details(invoice))
        
        # Recipient info
        story.extend(self._build_recipient_info(invoice))
        
        # Invoice items table
        story.extend(self._build_invoice_table(invoice))
        
        # Totals
        story.extend(self._build_totals_section(invoice))
        
        # Bank info and signature (only on last page)
        story.extend(self._build_footer_section(invoice))
        
        return story
    
    def _build_header(self) -> List:
        """Build company header section"""
        content = []
        
        # Company name
        company_name = Paragraph(
            app_config.name,
            self.styles['CustomHeader']
        )
        content.append(company_name)
        
        # Tagline
        tagline = Paragraph(
            f"<i>{app_config.company_tagline}</i>",
            self.styles['CompanyInfo']
        )
        content.append(tagline)
        
        # Office info
        office_info = f"{app_config.office_address_line1}<br/>{app_config.office_address_line2}<br/>Telp: {app_config.office_phone}"
        office = Paragraph(office_info, self.styles['CompanyInfo'])
        content.append(office)
        
        content.append(Spacer(1, 20))
        
        return content
    
    def _build_invoice_details(self, invoice: Invoice) -> List:
        """Build invoice details section"""
        content = []
        
        # Invoice title and details in table format
        invoice_data = [
            ['INVOICE', '', 'Jakarta, ' + format_date_long(invoice.invoice_date)],
            ['', '', ''],
            ['No. Invoice:', invoice.invoice_number, f'Halaman: 1 dari 1']
        ]
        
        invoice_table = Table(invoice_data, colWidths=[2*inch, 2*inch, 2*inch])
        invoice_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (0, 0), 14),
        ]))
        
        content.append(invoice_table)
        content.append(Spacer(1, 15))
        
        return content
    
    def _build_recipient_info(self, invoice: Invoice) -> List:
        """Build recipient company information"""
        content = []
        
        # Kepada section
        kepada = Paragraph("Kepada:", self.styles['InvoiceDetails'])
        content.append(kepada)
        
        # Company details
        company_info = f"<b>{invoice.company.company_name}</b><br/>"
        company_info += f"NPWP: {format_npwp_display(invoice.company.npwp)}<br/>"
        company_info += f"{invoice.company.address}"
        
        company_para = Paragraph(company_info, self.styles['InvoiceDetails'])
        content.append(company_para)
        
        content.append(Spacer(1, 15))
        
        return content
    
    def _build_invoice_table(self, invoice: Invoice) -> List:
        """Build invoice items table"""
        content = []
        
        # Table headers
        headers = ['No', 'Tanggal', 'Expatriat', 'Keterangan', 'Harga (Rp)']
        
        # Table data
        table_data = [headers]
        
        for i, line in enumerate(invoice.lines, 1):
            # Get TKA name
            tka_name = line.tka_worker.nama if line.tka_worker else "Unknown"
            
            # Get job description
            job_desc = line.custom_job_description or (
                line.job_description.job_description if line.job_description else ""
            )
            job_name = line.custom_job_name or (
                line.job_description.job_name if line.job_description else ""
            )
            
            # Format description with quantity if > 1
            keterangan = job_name
            if line.quantity > 1:
                keterangan += f" ({line.quantity}x)"
            if job_desc and job_desc != job_name:
                keterangan += f"\n{job_desc}"
            
            row = [
                str(line.baris),
                format_date_short(invoice.invoice_date),
                tka_name,
                keterangan,
                format_currency_idr(line.line_total, show_symbol=False)
            ]
            table_data.append(row)
        
        # Create table
        col_widths = [0.5*inch, 1*inch, 1.5*inch, 2.5*inch, 1.5*inch]
        invoice_table = Table(table_data, colWidths=col_widths)
        
        # Table styling
        table_style = [
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            
            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # No column
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Date column
            ('ALIGN', (4, 1), (4, -1), 'RIGHT'),   # Price column
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            
            # Borders
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
        ]
        
        invoice_table.setStyle(TableStyle(table_style))
        content.append(invoice_table)
        
        return content
    
    def _build_totals_section(self, invoice: Invoice) -> List:
        """Build totals section"""
        content = []
        
        content.append(Spacer(1, 15))
        
        # Totals table (right-aligned)
        totals_data = [
            ['', 'Subtotal:', format_currency_idr(invoice.subtotal, show_symbol=False)],
            ['', f'PPN {invoice.vat_percentage}%:', format_currency_idr(invoice.vat_amount, show_symbol=False)],
            ['', 'Total:', format_currency_idr(invoice.total_amount, show_symbol=False)],
        ]
        
        totals_table = Table(totals_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
        totals_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (1, -1), (-1, -1), 'Helvetica-Bold'),
            ('LINEABOVE', (1, -1), (-1, -1), 1, colors.black),
        ]))
        
        content.append(totals_table)
        
        # Amount in words
        content.append(Spacer(1, 10))
        amount_words = f"Terbilang: {format_currency_words(invoice.total_amount)}"
        words_para = Paragraph(amount_words, self.styles['InvoiceDetails'])
        content.append(words_para)
        
        return content
    
    def _build_footer_section(self, invoice: Invoice) -> List:
        """Build footer with bank info and signature"""
        content = []
        
        content.append(Spacer(1, 20))
        
        # Bank information
        if invoice.bank_account:
            bank_info = f"<b>Transfer ke:</b><br/>"
            bank_info += f"Bank: {invoice.bank_account.bank_name}<br/>"
            bank_info += f"No. Rekening: {invoice.bank_account.account_number}<br/>"
            bank_info += f"Atas Nama: {invoice.bank_account.account_name}"
            
            bank_para = Paragraph(bank_info, self.styles['InvoiceDetails'])
            content.append(bank_para)
            content.append(Spacer(1, 20))
        
        # Signature section
        signature_data = [
            ['', 'Hormat kami,'],
            ['', ''],
            ['', ''],
            ['', ''],
            ['', f'({invoice.creator.full_name})']
        ]
        
        signature_table = Table(signature_data, colWidths=[4*inch, 2*inch])
        signature_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ]))
        
        content.append(signature_table)
        
        return content

class ExcelExportService:
    """Service for exporting data to Excel format"""
    
    def __init__(self):
        self.font_name = 'Calibri'
        self.header_font = Font(name=self.font_name, size=12, bold=True)
        self.normal_font = Font(name=self.font_name, size=10)
        self.small_font = Font(name=self.font_name, size=9)
        
        # Colors
        self.header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_invoices_excel(self, invoices: List[Invoice], output_path: str = None) -> str:
        """
        Export multiple invoices to Excel
        
        Args:
            invoices: List of invoices to export
            output_path: Optional custom output path
            
        Returns:
            Path to generated Excel file
        """
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = safe_filename(f"Invoices_Export_{timestamp}.xlsx")
            output_dir = ensure_directory(export_config.export_directory)
            output_path = output_dir / filename
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create invoice summary sheet
        self._create_invoice_summary_sheet(wb, invoices)
        
        # Create invoice details sheet
        self._create_invoice_details_sheet(wb, invoices)
        
        # Save workbook
        wb.save(output_path)
        
        logger.info(f"Exported {len(invoices)} invoices to Excel: {output_path}")
        return str(output_path)
    
    def _create_invoice_summary_sheet(self, wb: openpyxl.Workbook, invoices: List[Invoice]):
        """Create invoice summary sheet"""
        ws = wb.create_sheet("Invoice Summary")
        
        # Headers
        headers = [
            'Invoice Number', 'Date', 'Company', 'NPWP', 'Status',
            'Subtotal', 'VAT Amount', 'Total Amount', 'Created By', 'Created Date'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row, invoice in enumerate(invoices, 2):
            data = [
                invoice.invoice_number,
                invoice.invoice_date,
                invoice.company.company_name,
                format_npwp_display(invoice.company.npwp),
                invoice.status.title(),
                float(invoice.subtotal),
                float(invoice.vat_amount),
                float(invoice.total_amount),
                invoice.creator.full_name,
                invoice.created_at.date() if invoice.created_at else None
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = self.normal_font
                cell.border = self.border
                
                # Format currency columns
                if col in [6, 7, 8]:  # Subtotal, VAT, Total columns
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_invoice_details_sheet(self, wb: openpyxl.Workbook, invoices: List[Invoice]):
        """Create invoice details sheet"""
        ws = wb.create_sheet("Invoice Details")
        
        # Headers
        headers = [
            'Invoice Number', 'Line No', 'TKA Name', 'Job Name', 
            'Job Description', 'Quantity', 'Unit Price', 'Line Total'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        current_row = 2
        for invoice in invoices:
            for line in invoice.lines:
                data = [
                    invoice.invoice_number,
                    line.baris,
                    line.tka_worker.nama if line.tka_worker else "Unknown",
                    line.custom_job_name or (line.job_description.job_name if line.job_description else ""),
                    line.custom_job_description or (line.job_description.job_description if line.job_description else ""),
                    line.quantity,
                    float(line.unit_price),
                    float(line.line_total)
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = value
                    cell.font = self.normal_font
                    cell.border = self.border
                    
                    # Format currency columns
                    if col in [7, 8]:  # Unit price, Line total
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal='right')
                
                current_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

class ExportService:
    """Main export service combining PDF and Excel functionality"""
    
    def __init__(self):
        self.pdf_service = PDFExportService()
        self.excel_service = ExcelExportService()
    
    def export_invoice_pdf(self, invoice_id: int, output_path: str = None) -> str:
        """Export single invoice to PDF"""
        return self.pdf_service.export_invoice_pdf(invoice_id, output_path)
    
    def export_invoices_excel(self, invoice_ids: List[int], output_path: str = None) -> str:
        """Export multiple invoices to Excel"""
        with InvoiceService() as service:
            invoices = []
            for invoice_id in invoice_ids:
                invoice = service.get_invoice(invoice_id)
                if invoice:
                    invoices.append(invoice)
        
        return self.excel_service.export_invoices_excel(invoices, output_path)
    
    def export_invoices_by_criteria(self, criteria: Dict[str, Any], 
                                  export_format: str = 'excel') -> str:
        """Export invoices based on criteria"""
        with InvoiceService() as service:
            # Get invoices based on criteria
            result = service.get_invoices_list(
                page=1, 
                per_page=10000,  # Large number to get all matching
                filters=criteria
            )
            
            invoice_ids = [inv['id'] for inv in result['invoices']]
            
            if export_format.lower() == 'pdf':
                # For PDF, export each invoice separately and return zip file path
                # This would require additional zip functionality
                raise NotImplementedError("Bulk PDF export not implemented yet")
            else:
                return self.export_invoices_excel(invoice_ids)

# Global export service instance
export_service = ExportService()

def export_invoice_pdf(invoice_id: int, output_path: str = None) -> str:
    """Convenience function to export invoice PDF"""
    return export_service.export_invoice_pdf(invoice_id, output_path)

def export_invoices_excel(invoice_ids: List[int], output_path: str = None) -> str:
    """Convenience function to export invoices to Excel"""
    return export_service.export_invoices_excel(invoice_ids, output_path)

if __name__ == "__main__":
    # Test export service
    print("Testing export service...")
    
    try:
        # Test PDF export (would need actual invoice ID)
        # pdf_path = export_invoice_pdf(1)
        # print(f"PDF exported to: {pdf_path}")
        
        # Test Excel export
        # excel_path = export_invoices_excel([1, 2, 3])
        # print(f"Excel exported to: {excel_path}")
        
        print("✅ Export service initialized successfully")
        
    except Exception as e:
        print(f"❌ Export service test failed: {e}")