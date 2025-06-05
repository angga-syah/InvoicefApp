"""
Invoice Management System - Import Service
Comprehensive service for importing data from Excel and CSV files
with validation, error reporting, and batch processing.
"""

import os
import csv
from pathlib import Path
from datetime import date, datetime
from decimal import Decimal
from typing import Dict, List, Optional, Any, Tuple, Union, BinaryIO
import logging

# Excel handling
import openpyxl
from openpyxl import load_workbook

# CSV handling
import pandas as pd

from models.database import (
    Company, TkaWorker, TkaFamilyMember, JobDescription, 
    Invoice, InvoiceLine, get_db_session
)
from services.invoice_service import InvoiceService
from utils.validators import (
    validate_company_data, validate_tka_worker_data, 
    validate_job_description_data, validate_invoice_data,
    validate_import_data, ValidationResult
)
from utils.helpers import safe_decimal, clean_string, normalize_name
from utils.formatters import generate_batch_id, safe_filename
from config import export_config

logger = logging.getLogger(__name__)

class ImportError(Exception):
    """Custom exception for import operations"""
    def __init__(self, message: str, row_number: int = None, field: str = None):
        self.message = message
        self.row_number = row_number
        self.field = field
        super().__init__(message)

class ImportResult:
    """Result of import operation"""
    def __init__(self):
        self.success_count = 0
        self.error_count = 0
        self.warning_count = 0
        self.errors: List[Dict] = []
        self.warnings: List[Dict] = []
        self.imported_ids: List[int] = []
        self.batch_id: str = generate_batch_id()
    
    def add_error(self, message: str, row_number: int = None, field: str = None):
        """Add import error"""
        self.error_count += 1
        self.errors.append({
            'message': message,
            'row_number': row_number,
            'field': field,
            'batch_id': self.batch_id
        })
    
    def add_warning(self, message: str, row_number: int = None, field: str = None):
        """Add import warning"""
        self.warning_count += 1
        self.warnings.append({
            'message': message,
            'row_number': row_number,
            'field': field,
            'batch_id': self.batch_id
        })
    
    def add_success(self, entity_id: int):
        """Add successful import"""
        self.success_count += 1
        self.imported_ids.append(entity_id)
    
    @property
    def total_processed(self) -> int:
        return self.success_count + self.error_count
    
    @property
    def is_successful(self) -> bool:
        return self.error_count == 0
    
    def get_summary(self) -> Dict[str, Any]:
        """Get import summary"""
        return {
            'batch_id': self.batch_id,
            'total_processed': self.total_processed,
            'success_count': self.success_count,
            'error_count': self.error_count,
            'warning_count': self.warning_count,
            'success_rate': (self.success_count / self.total_processed * 100) if self.total_processed > 0 else 0,
            'errors': self.errors,
            'warnings': self.warnings,
            'imported_ids': self.imported_ids
        }

class ExcelImportService:
    """Service for importing data from Excel files"""
    
    def __init__(self, session=None):
        self.session = session or get_db_session()
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.session:
            self.session.close()
    
    def read_excel_file(self, file_path: str, sheet_name: str = None) -> List[Dict[str, Any]]:
        """Read Excel file and return list of dictionaries"""
        try:
            workbook = load_workbook(file_path, data_only=True)
            
            # Get sheet
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ImportError(f"Sheet '{sheet_name}' not found in Excel file")
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            # Read headers from first row
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(clean_string(str(cell.value)).lower().replace(' ', '_'))
                else:
                    headers.append(f'column_{len(headers)}')
            
            # Read data rows
            data = []
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
                if not any(row):  # Skip empty rows
                    continue
                
                row_data = {'_row_number': row_num}
                for i, value in enumerate(row):
                    if i < len(headers):
                        if value is not None:
                            row_data[headers[i]] = clean_string(str(value)) if isinstance(value, str) else value
                        else:
                            row_data[headers[i]] = None
                
                data.append(row_data)
            
            return data
            
        except Exception as e:
            logger.error(f"Error reading Excel file {file_path}: {e}")
            raise ImportError(f"Failed to read Excel file: {str(e)}")
    
    def import_companies_excel(self, file_path: str, sheet_name: str = None) -> ImportResult:
        """Import companies from Excel file"""
        result = ImportResult()
        
        try:
            data = self.read_excel_file(file_path, sheet_name)
            
            for row_data in data:
                row_number = row_data.get('_row_number', 0)
                
                try:
                    # Map Excel columns to database fields
                    company_data = {
                        'company_name': row_data.get('company_name') or row_data.get('nama_perusahaan'),
                        'npwp': row_data.get('npwp'),
                        'idtku': row_data.get('idtku'),
                        'address': row_data.get('address') or row_data.get('alamat'),
                        'is_active': self._parse_boolean(row_data.get('is_active', True))
                    }
                    
                    # Validate data
                    validation = validate_company_data(company_data)
                    if not validation.is_valid:
                        for error in validation.errors:
                            result.add_error(error['message'], row_number, error['field'])
                        continue
                    
                    # Check for duplicates
                    existing = self.session.query(Company).filter(
                        (Company.npwp == company_data['npwp']) |
                        (Company.idtku == company_data['idtku'])
                    ).first()
                    
                    if existing:
                        result.add_warning(f"Company with NPWP/IDTKU already exists", row_number)
                        continue
                    
                    # Create company
                    company = Company(**company_data)
                    self.session.add(company)
                    self.session.flush()
                    
                    result.add_success(company.id)
                    
                except Exception as e:
                    result.add_error(f"Error processing row: {str(e)}", row_number)
                    continue
            
            if result.success_count > 0:
                self.session.commit()
            else:
                self.session.rollback()
            
            logger.info(f"Imported {result.success_count} companies from Excel file")
            return result
            
        except Exception as e:
            self.session.rollback()
            logger.error(f"Error importing companies from Excel: {e}")
            result.add_error(f"Import failed: {str(e)}")
            return result
    
    def import_tka_workers_excel(self, file_path: str, sheet_name: str = None) -> ImportResult:
        """Import TKA workers from Excel file"""
        result = ImportResult()
        
        try:
            data = self.read_excel_file(file_path, sheet_name)
            
            for row_data in data:
                row_number = row_data.get('_row_number', 0)
                
                try:
                    # Map Excel columns to database fields
                    tka_data = {
                        'nama': normalize_name(row_data.get('nama') or row_data.get('name', '')),
                        'passport': row_data.get('passport') or row_data.get('no_passport', ''),
                        'divisi': row_data.get('divisi') or row_data.get('division'),
                        'jenis_kelamin': self._normalize_gender(row_data.get('jenis_kelamin') or row_data.get('gender')),
                        'is_active': self._parse_boolean(row_data.get('is_active', True))
                    }
                    
                    # Validate data
                    validation = validate_tka_worker_data(tka_data)
                    if not validation.is_valid:
                        for error in validation.errors:
                            result.add_error(error['message'], row_number, error['field'])
                        continue
                    
                    # Check for duplicate passport
                    existing = self.session.query(TkaWorker).filter(
                        TkaWorker.passport == tka_data['passport']
                    ).first()
                    
                    if existing:
                        result.add_warning(f"TKA worker with passport {tka_data['passport']} already exists", row_number)
                        continue
                    
                    # Create TKA worker
                    tka_worker = TkaWorker(**tka_data)
                    self.session.add(tka_worker)
                    self.session.flush()
                    
                    result.add_success(tka_worker.id)
                    
                except Exception as e:
                    result.add_error(f"Error processing row: {str(e)}", row_number)
                    continue
            
            if result.success_count > 0:
                self.session.commit()
            else:
                self.session.rollback()
            
            logger.info(f"Imported {result.success_count} TKA workers from Excel file")
            return result
            
        except Exception as e:
            self.session.rollback()
            logger.error(f"Error importing TKA workers from Excel: {e}")
            result.add_error(f"Import failed: {str(e)}")
            return result
    
    def import_job_descriptions_excel(self, file_path: str, sheet_name: str = None) -> ImportResult:
        """Import job descriptions from Excel file"""
        result = ImportResult()
        
        try:
            data = self.read_excel_file(file_path, sheet_name)
            
            for row_data in data:
                row_number = row_data.get('_row_number', 0)
                
                try:
                    # Get company by NPWP or name
                    company_identifier = row_data.get('company_npwp') or row_data.get('company_name')
                    if not company_identifier:
                        result.add_error("Company identifier (NPWP or name) required", row_number, 'company')
                        continue
                    
                    company = self.session.query(Company).filter(
                        (Company.npwp == company_identifier) |
                        (Company.company_name.ilike(f'%{company_identifier}%'))
                    ).first()
                    
                    if not company:
                        result.add_error(f"Company not found: {company_identifier}", row_number, 'company')
                        continue
                    
                    # Map Excel columns to database fields
                    job_data = {
                        'company_id': company.id,
                        'job_name': row_data.get('job_name') or row_data.get('nama_pekerjaan'),
                        'job_description': row_data.get('job_description') or row_data.get('deskripsi_pekerjaan'),
                        'price': safe_decimal(row_data.get('price') or row_data.get('harga')),
                        'is_active': self._parse_boolean(row_data.get('is_active', True)),
                        'sort_order': int(row_data.get('sort_order', 0))
                    }
                    
                    # Validate data
                    validation = validate_job_description_data(job_data)
                    if not validation.is_valid:
                        for error in validation.errors:
                            result.add_error(error['message'], row_number, error['field'])
                        continue
                    
                    # Create job description
                    job_desc = JobDescription(**job_data)
                    self.session.add(job_desc)
                    self.session.flush()
                    
                    result.add_success(job_desc.id)
                    
                except Exception as e:
                    result.add_error(f"Error processing row: {str(e)}", row_number)
                    continue
            
            if result.success_count > 0:
                self.session.commit()
            else:
                self.session.rollback()
            
            logger.info(f"Imported {result.success_count} job descriptions from Excel file")
            return result
            
        except Exception as e:
            self.session.rollback()
            logger.error(f"Error importing job descriptions from Excel: {e}")
            result.add_error(f"Import failed: {str(e)}")
            return result
    
    def import_invoices_excel(self, file_path: str, user_id: int, sheet_name: str = None) -> ImportResult:
        """Import invoices from Excel file"""
        result = ImportResult()
        
        try:
            data = self.read_excel_file(file_path, sheet_name)
            
            # Group rows by invoice number
            invoice_groups = {}
            for row_data in data:
                invoice_number = row_data.get('invoice_number') or row_data.get('no_invoice')
                if not invoice_number:
                    result.add_error("Invoice number required", row_data.get('_row_number', 0))
                    continue
                
                if invoice_number not in invoice_groups:
                    invoice_groups[invoice_number] = []
                invoice_groups[invoice_number].append(row_data)
            
            # Process each invoice
            for invoice_number, rows in invoice_groups.items():
                try:
                    # Use first row for invoice header data
                    header_row = rows[0]
                    row_number = header_row.get('_row_number', 0)
                    
                    # Get company
                    company_identifier = header_row.get('company_npwp') or header_row.get('company_name')
                    company = self.session.query(Company).filter(
                        (Company.npwp == company_identifier) |
                        (Company.company_name.ilike(f'%{company_identifier}%'))
                    ).first()
                    
                    if not company:
                        result.add_error(f"Company not found: {company_identifier}", row_number)
                        continue
                    
                    # Prepare invoice data
                    invoice_data = {
                        'invoice_number': invoice_number,
                        'company_id': company.id,
                        'invoice_date': self._parse_date(header_row.get('invoice_date') or header_row.get('tanggal_invoice')),
                        'vat_percentage': safe_decimal(header_row.get('vat_percentage', 11)),
                        'status': header_row.get('status', 'draft'),
                        'notes': header_row.get('notes') or header_row.get('catatan'),
                        'imported_from': 'excel',
                        'import_batch_id': result.batch_id
                    }
                    
                    # Prepare line items
                    line_items = []
                    for i, row in enumerate(rows):
                        # Get TKA worker
                        tka_identifier = row.get('tka_passport') or row.get('tka_name')
                        tka_worker = self.session.query(TkaWorker).filter(
                            (TkaWorker.passport == tka_identifier) |
                            (TkaWorker.nama.ilike(f'%{tka_identifier}%'))
                        ).first()
                        
                        if not tka_worker:
                            result.add_error(f"TKA worker not found: {tka_identifier}", row.get('_row_number', 0))
                            continue
                        
                        # Get job description
                        job_name = row.get('job_name') or row.get('nama_pekerjaan')
                        job_desc = self.session.query(JobDescription).filter(
                            JobDescription.company_id == company.id,
                            JobDescription.job_name.ilike(f'%{job_name}%')
                        ).first()
                        
                        if not job_desc:
                            result.add_error(f"Job description not found: {job_name}", row.get('_row_number', 0))
                            continue
                        
                        line_data = {
                            'tka_id': tka_worker.id,
                            'job_description_id': job_desc.id,
                            'custom_price': safe_decimal(row.get('custom_price')) if row.get('custom_price') else None,
                            'quantity': int(row.get('quantity', 1)),
                            'baris': int(row.get('baris', i + 1))
                        }
                        line_items.append(line_data)
                    
                    if not line_items:
                        result.add_error(f"No valid line items for invoice {invoice_number}", row_number)
                        continue
                    
                    # Create invoice using service
                    with InvoiceService(self.session) as invoice_service:
                        invoice, validation = invoice_service.create_invoice(invoice_data, line_items, user_id)
                        
                        if validation.is_valid:
                            result.add_success(invoice.id)
                        else:
                            for error in validation.errors:
                                result.add_error(error['message'], row_number, error.get('field'))
                
                except Exception as e:
                    result.add_error(f"Error processing invoice {invoice_number}: {str(e)}", row_number)
                    continue
            
            logger.info(f"Imported {result.success_count} invoices from Excel file")
            return result
            
        except Exception as e:
            logger.error(f"Error importing invoices from Excel: {e}")
            result.add_error(f"Import failed: {str(e)}")
            return result
    
    def _parse_boolean(self, value: Any) -> bool:
        """Parse boolean value from various formats"""
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ('true', '1', 'yes', 'ya', 'active', 'aktif')
        if isinstance(value, (int, float)):
            return bool(value)
        return True  # Default to True
    
    def _normalize_gender(self, gender: str) -> str:
        """Normalize gender values"""
        if not gender:
            return 'Laki-laki'
        
        gender_lower = gender.lower()
        if gender_lower in ('male', 'laki-laki', 'laki', 'l', 'm'):
            return 'Laki-laki'
        elif gender_lower in ('female', 'perempuan', 'wanita', 'p', 'f'):
            return 'Perempuan'
        else:
            return 'Laki-laki'  # Default
    
    def _parse_date(self, value: Any) -> date:
        """Parse date from various formats"""
        if isinstance(value, date):
            return value
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, str):
            # Try common date formats
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        
        return date.today()  # Default to today

class CSVImportService:
    """Service for importing data from CSV files"""
    
    def __init__(self, session=None):
        self.session = session or get_db_session()
        self.excel_service = ExcelImportService(session)
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.session:
            self.session.close()
    
    def read_csv_file(self, file_path: str, encoding: str = 'utf-8') -> List[Dict[str, Any]]:
        """Read CSV file and return list of dictionaries"""
        try:
            # Try different encodings if UTF-8 fails
            encodings = [encoding, 'utf-8', 'latin-1', 'cp1252']
            
            for enc in encodings:
                try:
                    df = pd.read_csv(file_path, encoding=enc)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise ImportError("Could not decode CSV file with any supported encoding")
            
            # Clean column names
            df.columns = [clean_string(str(col)).lower().replace(' ', '_') for col in df.columns]
            
            # Convert to list of dictionaries
            data = []
            for index, row in df.iterrows():
                row_data = {'_row_number': index + 2}  # +2 because pandas is 0-based and we skip header
                for col, value in row.items():
                    if pd.notna(value):
                        row_data[col] = clean_string(str(value)) if isinstance(value, str) else value
                    else:
                        row_data[col] = None
                data.append(row_data)
            
            return data
            
        except Exception as e:
            logger.error(f"Error reading CSV file {file_path}: {e}")
            raise ImportError(f"Failed to read CSV file: {str(e)}")
    
    def import_companies_csv(self, file_path: str, encoding: str = 'utf-8') -> ImportResult:
        """Import companies from CSV file"""
        # Convert CSV to Excel-like format and use Excel service
        data = self.read_csv_file(file_path, encoding)
        return self._import_companies_from_data(data)
    
    def import_tka_workers_csv(self, file_path: str, encoding: str = 'utf-8') -> ImportResult:
        """Import TKA workers from CSV file"""
        data = self.read_csv_file(file_path, encoding)
        return self._import_tka_workers_from_data(data)
    
    def import_job_descriptions_csv(self, file_path: str, encoding: str = 'utf-8') -> ImportResult:
        """Import job descriptions from CSV file"""
        data = self.read_csv_file(file_path, encoding)
        return self._import_job_descriptions_from_data(data)
    
    def _import_companies_from_data(self, data: List[Dict]) -> ImportResult:
        """Import companies from data list"""
        result = ImportResult()
        
        for row_data in data:
            row_number = row_data.get('_row_number', 0)
            
            try:
                company_data = {
                    'company_name': row_data.get('company_name') or row_data.get('nama_perusahaan'),
                    'npwp': row_data.get('npwp'),
                    'idtku': row_data.get('idtku'),
                    'address': row_data.get('address') or row_data.get('alamat'),
                    'is_active': self.excel_service._parse_boolean(row_data.get('is_active', True))
                }
                
                validation = validate_company_data(company_data)
                if not validation.is_valid:
                    for error in validation.errors:
                        result.add_error(error['message'], row_number, error['field'])
                    continue
                
                existing = self.session.query(Company).filter(
                    (Company.npwp == company_data['npwp']) |
                    (Company.idtku == company_data['idtku'])
                ).first()
                
                if existing:
                    result.add_warning(f"Company with NPWP/IDTKU already exists", row_number)
                    continue
                
                company = Company(**company_data)
                self.session.add(company)
                self.session.flush()
                
                result.add_success(company.id)
                
            except Exception as e:
                result.add_error(f"Error processing row: {str(e)}", row_number)
                continue
        
        if result.success_count > 0:
            self.session.commit()
        else:
            self.session.rollback()
        
        return result
    
    def _import_tka_workers_from_data(self, data: List[Dict]) -> ImportResult:
        """Import TKA workers from data list"""
        result = ImportResult()
        
        for row_data in data:
            row_number = row_data.get('_row_number', 0)
            
            try:
                tka_data = {
                    'nama': normalize_name(row_data.get('nama') or row_data.get('name', '')),
                    'passport': row_data.get('passport') or row_data.get('no_passport', ''),
                    'divisi': row_data.get('divisi') or row_data.get('division'),
                    'jenis_kelamin': self.excel_service._normalize_gender(row_data.get('jenis_kelamin') or row_data.get('gender')),
                    'is_active': self.excel_service._parse_boolean(row_data.get('is_active', True))
                }
                
                validation = validate_tka_worker_data(tka_data)
                if not validation.is_valid:
                    for error in validation.errors:
                        result.add_error(error['message'], row_number, error['field'])
                    continue
                
                existing = self.session.query(TkaWorker).filter(
                    TkaWorker.passport == tka_data['passport']
                ).first()
                
                if existing:
                    result.add_warning(f"TKA worker with passport {tka_data['passport']} already exists", row_number)
                    continue
                
                tka_worker = TkaWorker(**tka_data)
                self.session.add(tka_worker)
                self.session.flush()
                
                result.add_success(tka_worker.id)
                
            except Exception as e:
                result.add_error(f"Error processing row: {str(e)}", row_number)
                continue
        
        if result.success_count > 0:
            self.session.commit()
        else:
            self.session.rollback()
        
        return result
    
    def _import_job_descriptions_from_data(self, data: List[Dict]) -> ImportResult:
        """Import job descriptions from data list"""
        result = ImportResult()
        
        for row_data in data:
            row_number = row_data.get('_row_number', 0)
            
            try:
                company_identifier = row_data.get('company_npwp') or row_data.get('company_name')
                if not company_identifier:
                    result.add_error("Company identifier (NPWP or name) required", row_number, 'company')
                    continue
                
                company = self.session.query(Company).filter(
                    (Company.npwp == company_identifier) |
                    (Company.company_name.ilike(f'%{company_identifier}%'))
                ).first()
                
                if not company:
                    result.add_error(f"Company not found: {company_identifier}", row_number, 'company')
                    continue
                
                job_data = {
                    'company_id': company.id,
                    'job_name': row_data.get('job_name') or row_data.get('nama_pekerjaan'),
                    'job_description': row_data.get('job_description') or row_data.get('deskripsi_pekerjaan'),
                    'price': safe_decimal(row_data.get('price') or row_data.get('harga')),
                    'is_active': self.excel_service._parse_boolean(row_data.get('is_active', True)),
                    'sort_order': int(row_data.get('sort_order', 0))
                }
                
                validation = validate_job_description_data(job_data)
                if not validation.is_valid:
                    for error in validation.errors:
                        result.add_error(error['message'], row_number, error['field'])
                    continue
                
                job_desc = JobDescription(**job_data)
                self.session.add(job_desc)
                self.session.flush()
                
                result.add_success(job_desc.id)
                
            except Exception as e:
                result.add_error(f"Error processing row: {str(e)}", row_number)
                continue
        
        if result.success_count > 0:
            self.session.commit()
        else:
            self.session.rollback()
        
        return result

class ImportService:
    """Main import service combining Excel and CSV functionality"""
    
    def __init__(self, session=None):
        self.session = session or get_db_session()
        self.excel_service = ExcelImportService(session)
        self.csv_service = CSVImportService(session)
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.session:
            self.session.close()
    
    def import_file(self, file_path: str, entity_type: str, user_id: int = None, 
                   file_format: str = None, **kwargs) -> ImportResult:
        """
        Import data from file (auto-detect format)
        
        Args:
            file_path: Path to file
            entity_type: Type of entity (companies, tka_workers, job_descriptions, invoices)
            user_id: User performing import (required for invoices)
            file_format: Force specific format (excel, csv)
            **kwargs: Additional parameters for import
        """
        # Auto-detect format if not specified
        if not file_format:
            file_ext = Path(file_path).suffix.lower()
            if file_ext in ['.xlsx', '.xls']:
                file_format = 'excel'
            elif file_ext == '.csv':
                file_format = 'csv'
            else:
                raise ImportError(f"Unsupported file format: {file_ext}")
        
        # Route to appropriate service
        if file_format == 'excel':
            return self._import_excel(file_path, entity_type, user_id, **kwargs)
        elif file_format == 'csv':
            return self._import_csv(file_path, entity_type, user_id, **kwargs)
        else:
            raise ImportError(f"Unsupported file format: {file_format}")
    
    def _import_excel(self, file_path: str, entity_type: str, user_id: int = None, **kwargs) -> ImportResult:
        """Import from Excel file"""
        sheet_name = kwargs.get('sheet_name')
        
        if entity_type == 'companies':
            return self.excel_service.import_companies_excel(file_path, sheet_name)
        elif entity_type == 'tka_workers':
            return self.excel_service.import_tka_workers_excel(file_path, sheet_name)
        elif entity_type == 'job_descriptions':
            return self.excel_service.import_job_descriptions_excel(file_path, sheet_name)
        elif entity_type == 'invoices':
            if not user_id:
                raise ImportError("User ID required for invoice import")
            return self.excel_service.import_invoices_excel(file_path, user_id, sheet_name)
        else:
            raise ImportError(f"Unsupported entity type: {entity_type}")
    
    def _import_csv(self, file_path: str, entity_type: str, user_id: int = None, **kwargs) -> ImportResult:
        """Import from CSV file"""
        encoding = kwargs.get('encoding', 'utf-8')
        
        if entity_type == 'companies':
            return self.csv_service.import_companies_csv(file_path, encoding)
        elif entity_type == 'tka_workers':
            return self.csv_service.import_tka_workers_csv(file_path, encoding)
        elif entity_type == 'job_descriptions':
            return self.csv_service.import_job_descriptions_csv(file_path, encoding)
        elif entity_type == 'invoices':
            # CSV invoice import would need to be implemented similar to Excel
            raise ImportError("CSV invoice import not yet implemented")
        else:
            raise ImportError(f"Unsupported entity type: {entity_type}")

# Global import service instance
import_service = ImportService()

def import_data_file(file_path: str, entity_type: str, user_id: int = None, **kwargs) -> ImportResult:
    """Convenience function to import data file"""
    with ImportService() as service:
        return service.import_file(file_path, entity_type, user_id, **kwargs)

if __name__ == "__main__":
    # Test import service
    print("Testing import service...")
    
    try:
        # Test Excel reading (would need actual file)
        # result = import_data_file('test_companies.xlsx', 'companies')
        # print(f"Import result: {result.get_summary()}")
        
        print("✅ Import service initialized successfully")
        
    except Exception as e:
        print(f"❌ Import service test failed: {e}")